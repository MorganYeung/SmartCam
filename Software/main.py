# SDC GUI v1
import datetime
import math
import random
import sys
import time
import datetime
from datetime import date

import kivy
import kivy.core.text
import cv2
from kivy.app import App
from kivy.base import EventLoop
from kivy.uix.image import Image
from kivy.graphics.texture import Texture
from kivy.core.window import Window
#import matplotlib as mp1
#mp1.use('Agg')
#import matplotlib.pyplot as plt
from kivy.app import App
from kivy.clock import Clock
from kivy.config import Config
from kivy.core.window import Window
#from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
from kivy.properties import NumericProperty, ObjectProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.checkbox import CheckBox
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.screenmanager import Screen, ScreenManager
from kivy.uix.spinner import Spinner
from kivy.uix.textinput import TextInput
from kivy.uix.widget import Widget
from openpyxl import Workbook, load_workbook

# Screen Configuration
Config.read('config.ini')
Config.set('graphics', 'width', 1920)
Config.set('graphics', 'height', 1080)
Config.set('graphics', 'fullscreen', 0)
Config.set('graphics', 'borderless',0)
Config.set('kivy','keyboard_mode', 'dock')
Config.write()

#Excel Documents
#Results Logging
#test_data_book = Workbook()
#test_data = test_data_book.active

#Limit File to compare Results with
#limits = load_workbook('Limits.xlsx')
#limitsL = limits.active

#Constants
class StartScreen(Screen):
    def exit(self):
        sys.exit()

class KivyCamera(Image):

    def __init__(self, **kwargs):
        super(KivyCamera, self).__init__(**kwargs)
        self.capture = None

    def start(self, capture, fps=30):
        self.capture = capture
        Clock.schedule_interval(self.update, 1.0 / fps)

    def stop(self):
        Clock.unschedule_interval(self.update)
        self.capture = None

    def update(self, dt):
        return_value, frame = self.capture.read()
        if return_value:
            texture = self.texture
            w, h = frame.shape[1], frame.shape[0]
            if not texture or texture.width != w or texture.height != h:
                self.texture = texture = Texture.create(size=(w, h))
                texture.flip_vertical()
            texture.blit_buffer(frame.tobytes(), colorfmt='bgr')
            self.canvas.ask_update()


capture = None

class MainScreen(Screen):
    def init_qrtest(self):
        pass

    def dostart(self, *largs):
        global capture
        capture = cv2.VideoCapture(0)
        self.ids.qrcam.start(capture)

    def doexit(self):
        global capture
        if capture != None:
            capture.release()
            capture = None
        EventLoop.close()
    def exit(self):
        sys.exit()


class DataScreen(Screen):
    def exit(self):
        sys.exit()
#Popup classes are empty, see lat.kv
class WarningPop(Popup):
    pass
class PassPop(Popup):
    pass
class FailPop(Popup):
    pass


class Manager(ScreenManager):
    StartScreen = ObjectProperty(None)
    MainScreen = ObjectProperty(None)
    DataScreen = ObjectProperty(None)

class SDCApp(App):
    loss_test = 0
    test = 0
    loss_test_value = 0
    laser_test_count = 0
    attenuation_test_count = 0
    found_flag = 0
    pass_flag = 0
    min_pass_flag = 0
    max_pass_flag = 0
    laser_power = []
    DAC = None
    measurement = None
    current = None
    dac_25mW= 0
    dac_50mW= 0
    dac_75mW= 0
    baseline1 = 25
    baseline2 = 50
    baseline3 = 75
    test_point1 = 0
    test_point2 = 0
    test_point3 = 0
    def build(self):
        m = Manager()
        return m

if __name__ == '__main__':
    SDCApp().run()
